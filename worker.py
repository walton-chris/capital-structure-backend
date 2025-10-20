import os, io, json, uuid, logging, hashlib, re
from typing import Any, Dict, List, Optional
from collections import defaultdict
import openpyxl
from pydantic import BaseModel, Field, ValidationError, NonNegativeFloat, field_validator
from redis import Redis
import requests
from openai import OpenAI

# --- Configuration & Logging ---
logging.basicConfig(level=os.getenv("LOG_LEVEL", "INFO").upper(), format="%(asctime)s %(levelname)s [worker] - %(message)s")
logger = logging.getLogger("worker")

REDIS_URL = os.getenv("REDIS_URL")
if not REDIS_URL: logger.critical("FATAL: REDIS_URL environment variable is not set.")
redis = Redis.from_url(REDIS_URL)

OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")
client: Optional[OpenAI] = OpenAI(api_key=OPENAI_API_KEY) if OPENAI_API_KEY else None

# --- Pydantic Models (must be defined in worker too) ---
class Security(BaseModel):
    name: str; shares_outstanding: NonNegativeFloat; original_investment_per_share: NonNegativeFloat
    liquidation_preference_multiple: NonNegativeFloat; seniority: Optional[int] = Field(default=None, ge=0, le=10)
    is_participating: bool; participation_cap_multiple: NonNegativeFloat; cumulative_dividend_rate: NonNegativeFloat
    years_since_issuance: NonNegativeFloat
    @field_validator("name")
    def _non_empty(cls, v: str) -> str:
        if not v or not v.strip(): raise ValueError("Security 'name' cannot be empty.")
        return v.strip()
class CapitalStructureInput(BaseModel): securities: List[Security]; total_option_pool_shares: NonNegativeFloat

# --- Deterministic Parsing Logic ---
TOTALS_RE = re.compile(r"\b(total|subtotal|grand\s+total)\b", re.IGNORECASE)

class Anonymizer:
    # ... (Anonymizer class from previous step) ...
    def __init__(self):
        self.map: Dict[str,str] = {}; self.person_counter = 1; self.entity_counter = 1
        self.entity_re = re.compile(r"(LLC|Inc|LP|FBO|Capital|Partners|Fund|Trust|Ventures)", re.IGNORECASE)
    def maybe(self, name: Any) -> Any:
        if not isinstance(name, str): return name
        key = name.strip()
        if not key: return name
        if key in self.map: return self.map[key]
        is_entity = bool(self.entity_re.search(key)) or key.isupper()
        ph = f"Entity-{self.entity_counter}" if is_entity else f"Person-{self.person_counter}"
        if is_entity: self.entity_counter += 1
        else: self.person_counter += 1
        self.map[key] = ph
        return ph

def normalize_header(h: Any) -> str: return re.sub(r"\s+", " ", str(h or "")).strip().lower()

def as_float(x: Any, default: float=0.0) -> float:
    try: return float(str(x).replace(",","").replace("$",""))
    except (ValueError, TypeError): return default

def parse_sheet(sheet, header_aliases, anonymizer, anonymize_headers):
    # ... (deterministic parse_sheet function from expert feedback) ...
    rows = list(sheet.iter_rows(values_only=True));
    if not rows: return []
    header_idx, best_hits = -1, -1
    for i, row in enumerate(rows[:20]):
        norm = [normalize_header(c) for c in row]
        hits = sum(1 for h in norm if h in header_aliases)
        if hits > best_hits and hits >= 2: header_idx, best_hits = i, hits
    if header_idx < 0: return []
    headers = [normalize_header(c) for c in rows[header_idx]]
    header_map = {h: header_aliases[h] for h in headers if h in header_aliases}
    anon_idx = {i for i, h in enumerate(headers) if h in anonymize_headers}
    data: List[Dict[str,Any]] = []
    for row in rows[header_idx+1:]:
        if not any(row) or any(isinstance(c, str) and TOTALS_RE.search(c) for c in row if c): continue
        rec: Dict[str,Any] = {}
        for i, cell in enumerate(row):
            if i >= len(headers): break
            src_h = headers[i]
            if src_h not in header_map: continue
            dst = header_map[src_h]
            val = anonymizer.maybe(cell) if i in anon_idx else cell
            rec[dst] = val
        if rec: data.append(rec)
    return data

# --- The Main Job Function ---
def run_extract_job(file_path: str, file_sha256: str, callback_url: Optional[str] = None) -> Dict[str, Any]:
    rid = str(uuid.uuid4())[:8]
    logger.info(f"[rid={rid}] Starting job for file: {file_path}")
    cache_key = f"result:{file_sha256}"
    
    try:
        with open(file_path, "rb") as f:
            file_bytes = f.read()
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as e:
        logger.exception(f"[rid={rid}] Worker failed to load workbook.")
        raise IOError(f"Cannot load workbook: {e}")

    anonymizer = Anonymizer()
    cap_sheet, opt_sheet = None, None
    for sh in wb.worksheets:
        t = (sh.title or "").lower()
        if not cap_sheet and ("detailed cap" in t or "cap table" in t): cap_sheet = sh
        if not opt_sheet and ("option" in t or "grant" in t): opt_sheet = sh
    if not cap_sheet: cap_sheet = wb.worksheets[0]

    cap_aliases = {"class":"name", "class of stock":"name", "shares outstanding":"shares", "total outstanding":"shares", "price per share":"price"}
    opt_aliases = {"optionholder":"holder", "holder":"holder", "name":"holder", "options granted":"shares", "shares":"shares", "exercise price":"price", "strike price":"price"}

    # NOTE: This implementation is fully deterministic and does not use the LLM.
    # The LLM "assist" can be added back here if deterministic parsing proves insufficient.
    cap_rows = parse_sheet(cap_sheet, cap_aliases, anonymizer, anonymize_headers=("stakeholder",))
    opt_rows = parse_sheet(opt_sheet, opt_aliases, anonymizer, anonymize_headers=("optionholder", "holder", "name")) if opt_sheet else []

    securities: List[Dict[str, Any]] = []
    # This logic assumes pivoted table for cap sheet (like in screenshots)
    if cap_rows and isinstance(cap_rows[0].get('name'), str): # Check if parsing was row-based
        # Row-based parsing logic... (add if needed)
        pass # Placeholder for now
    elif cap_sheet: # Pivoted table parsing
        headers = [normalize_header(c.value) for c in cap_sheet[1]]
        total_row = None
        for row in cap_sheet.iter_rows(values_only=True):
            if isinstance(row[0], str) and "total shares outstanding" in row[0].lower():
                total_row = row
                break
        if total_row:
            for i, header in enumerate(headers):
                if header in cap_aliases and cap_aliases[header] == 'name':
                    is_pref = "preferred" in header
                    securities.append({
                        "name": str(cap_sheet.cell(row=1, column=i+1).value),
                        "shares_outstanding": as_float(total_row[i]),
                        "original_investment_per_share": 0.0, # This info is often elsewhere
                        "liquidation_preference_multiple": 1.0 if is_pref else 0.0,
                        "seniority": 1 if is_pref else None, "is_participating": False,
                        "participation_cap_multiple": 0.0, "cumulative_dividend_rate": 0.0, "years_since_issuance": 0.0
                    })

    options_by_px = defaultdict(float)
    for r in opt_rows:
        px = as_float(r.get("price"), 0.0)
        sh = as_float(r.get("shares"), 0.0)
        if px > 0 and sh > 0:
            options_by_px[px] += sh
    for px, total in options_by_px.items():
        securities.append({
            "name": f"Options at ${px:.2f} Exercise Price", "shares_outstanding": total,
            "original_investment_per_share": 0.0, "liquidation_preference_multiple": 0.0,
            "seniority": None, "is_participating": False, "participation_cap_multiple": 0.0,
            "cumulative_dividend_rate": 0.0, "years_since_issuance": 0.0
        })

    result_obj = {"securities": securities, "total_option_pool_shares": 0.0}
    validated = CapitalStructureInput.model_validate(result_obj)
    result = json.loads(validated.model_dump_json())
    
    redis.setex(cache_key, 86400, json.dumps(result))
    if callback_url:
        try:
            requests.post(callback_url, json={"status": "succeeded", "result": result}, timeout=5)
        except Exception as e:
            logger.warning(f"[rid={rid}] Webhook failed: {e}")
    
    try:
        os.remove(file_path)
        logger.info(f"[rid={rid}] Cleaned up temp file: {file_path}")
    except OSError as e:
        logger.warning(f"[rid={rid}] Could not clean up temp file (may have been deleted): {e}")

    logger.info(f"[rid={rid}] Job completed successfully.")
    return result
