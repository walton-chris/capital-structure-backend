import os
import base64
import uuid
import json
import io
import re
from typing import List, Optional, Dict, Any
from collections import defaultdict
from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
import openai
import openpyxl

# Initialize FastAPI app
app = FastAPI(
    title="Capital Structure API",
    version="2.0.0"
)

# Configure CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure OpenAI
openai.api_key = os.getenv("OPENAI_API_KEY")

# In-memory file storage
file_storage = {}

# Pydantic models
class Security(BaseModel):
    name: str
    shares_outstanding: float
    original_investment_per_share: float
    liquidation_preference_multiple: float
    seniority: Optional[int] = None
    is_participating: bool
    participation_cap_multiple: float
    cumulative_dividend_rate: float
    years_since_issuance: float

class CapitalStructureInput(BaseModel):
    securities: List[Security]
    total_option_pool_shares: float

class FileUploadRequest(BaseModel):
    file_content: str
    file_name: str

class DocumentExtractRequest(BaseModel):
    file_id: str

class DocumentUploadResponse(BaseModel):
    file_id: str
    file_name: str
    message: str
    file_size_bytes: int

# Excel parsing functions
def clean_security_name(name: str) -> str:
    """Remove abbreviations in parentheses and clean up security names"""
    if not name:
        return ""
    # Remove content in parentheses like (CS), (PS), (PA), etc.
    cleaned = re.sub(r'\s*\([^)]*\)\s*', '', str(name))
    # Remove conversion ratio text
    cleaned = re.sub(r'\s*\d+:\d+\s*Conversion Ratio', '', cleaned, flags=re.IGNORECASE)
    return cleaned.strip()

def is_conversion_ratio_column(header: str) -> bool:
    """Check if a column header indicates it's a conversion ratio column"""
    if not header:
        return False
    header_lower = str(header).lower()
    return 'conversion ratio' in header_lower or '1:1' in header_lower

def parse_excel_cap_table(file_bytes: bytes) -> Dict[str, Any]:
    """Parse Excel file and extract structured cap table data"""
    try:
        workbook = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        
        # Find sheets by content
        cap_table_sheet = None
        option_ledger_sheet = None
        
        for sheet in workbook.worksheets:
            sheet_name_lower = sheet.title.lower()
            
            # Check first few rows for content indicators
            first_rows_text = ' '.join([
                str(cell or '') 
                for row in sheet.iter_rows(min_row=1, max_row=5, values_only=True)
                for cell in row
            ]).lower()
            
            if ('cap table' in sheet_name_lower or 'detailed' in sheet_name_lower or 
                ('common' in first_rows_text and 'preferred' in first_rows_text)):
                cap_table_sheet = sheet
            elif 'option' in sheet_name_lower or 'grant' in sheet_name_lower:
                option_ledger_sheet = sheet
        
        # Use first sheet as cap table if none found
        if cap_table_sheet is None:
            cap_table_sheet = workbook.worksheets[0]
        
        # Extract data
        cap_table_data = extract_cap_table_structure(cap_table_sheet)
        option_data = extract_option_ledger(option_ledger_sheet) if option_ledger_sheet else []
        
        return {
            "cap_table": cap_table_data,
            "option_ledger": option_data,
            "source": "excel"
        }
    
    except Exception as e:
        raise Exception(f"Failed to parse Excel file: {str(e)}")

def extract_cap_table_structure(sheet) -> Dict[str, Any]:
    """Extract structured data from cap table sheet"""
    data = {
        "headers": [],
        "raw_headers": [],
        "totals": {},
        "prices": {},
        "options_outstanding": None,
        "option_pool_available": None
    }
    
    # Convert sheet to list for easier processing
    all_rows = list(sheet.iter_rows(values_only=True))
    
    if len(all_rows) == 0:
        return data
    
    # Find header row (contains "Name" or "Stakeholder" + security class names)
    header_row_idx = None
    for idx, row in enumerate(all_rows[:15]):
        row_str = ' '.join([str(cell) for cell in row if cell]).lower()
        if ('common' in row_str and ('series' in row_str or 'preferred' in row_str)) or \
           ('name' in row_str and ('common' in row_str or 'stock' in row_str)):
            header_row_idx = idx
            data["raw_headers"] = list(row)
            
            # Process headers: clean names and skip conversion ratio columns
            cleaned_headers = []
            skip_next = False
            
            for i, header in enumerate(row):
                if skip_next:
                    skip_next = False
                    cleaned_headers.append(None)  # Placeholder for conversion column
                    continue
                
                if header and is_conversion_ratio_column(str(header)):
                    cleaned_headers.append(None)  # Mark as conversion column
                elif header:
                    clean_name = clean_security_name(header)
                    cleaned_headers.append(clean_name)
                    
                    # Check if next column is conversion ratio for this security
                    if i + 1 < len(row) and row[i + 1]:
                        next_header = str(row[i + 1])
                        if is_conversion_ratio_column(next_header) and clean_name in next_header:
                            skip_next = True
                else:
                    cleaned_headers.append(None)
            
            data["headers"] = cleaned_headers
            break
    
    if header_row_idx is None:
        return data
    
    # Search bottom 30 rows for summary data
    start_search = max(header_row_idx + 1, len(all_rows) - 30)
    
    print(f"DEBUG: Searching rows {start_search} to {len(all_rows)} for summary data")
    
    for idx in range(start_search, len(all_rows)):
        row = all_rows[idx]
        if not row or not row[0]:
            continue
        
        row_label = str(row[0]).strip().lower()
        
        # Debug: Print row labels we're checking
        if any(keyword in row_label for keyword in ['fully', 'total', 'price', 'option', 'available']):
            print(f"  Row {idx}: '{row_label}'")
        
        # Extract shares outstanding totals
        if 'fully diluted shares' in row_label or 'total shares outstanding' in row_label:
            print(f"    FOUND TOTALS ROW: '{row_label}'")
            for col_idx in range(1, min(len(row), len(data["headers"]))):
                cell_value = row[col_idx]
                header = data["headers"][col_idx]
                
                # Skip conversion columns and null headers
                if header is None:
                    continue
                
                # Check if this is actually a shares count (not a percentage)
                if cell_value and isinstance(cell_value, (int, float)) and cell_value > 1000:
                    print(f"      {header}: {cell_value}")
                    data["totals"][header] = float(cell_value)
        
        # Extract price per share
        elif 'price per share' in row_label:
            print(f"    FOUND PRICES ROW: '{row_label}'")
            for col_idx in range(1, min(len(row), len(data["headers"]))):
                cell_value = row[col_idx]
                header = data["headers"][col_idx]
                
                # Skip conversion columns and null headers
                if header is None or not cell_value:
                    continue
                
                # Clean and parse price
                price_str = str(cell_value).replace('
        
        # Extract options outstanding
        elif 'option' in row_label and 'outstanding' in row_label and 'issued' in row_label:
            print(f"    FOUND OPTIONS OUTSTANDING ROW: '{row_label}'")
            for cell_value in row[1:]:
                if cell_value and isinstance(cell_value, (int, float)) and cell_value > 0:
                    print(f"      Options outstanding: {cell_value}")
                    data["options_outstanding"] = float(cell_value)
                    break
        
        # Extract option pool available
        elif 'available for issuance' in row_label or ('shares available' in row_label and 'plan' in row_label):
            print(f"    FOUND OPTION POOL AVAILABLE ROW: '{row_label}'")
            for cell_value in row[1:]:
                if cell_value and isinstance(cell_value, (int, float)) and cell_value > 0:
                    print(f"      Option pool available: {cell_value}")
                    data["option_pool_available"] = float(cell_value)
                    break
    
    return data

def extract_option_ledger(sheet) -> List[Dict[str, Any]]:
    """Extract option grants from option ledger sheet"""
    options = []
    
    # Find header row
    header_row_idx = None
    headers = {}
    
    for idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=15, values_only=True)):
        row_text = ' '.join([str(cell or '').lower() for cell in row])
        
        # Look for key indicators of header row
        if 'exercise' in row_text and 'price' in row_text:
            header_row_idx = idx + 1
            
            # Map column names to indices
            for col_idx, cell in enumerate(row):
                if not cell:
                    continue
                
                col_name = str(cell).lower().strip()
                
                # Look for key columns
                if 'outstanding' in col_name and 'option' in col_name:
                    headers['options_outstanding'] = col_idx
                elif 'granted' in col_name and 'option' in col_name:
                    headers['options_granted'] = col_idx
                elif 'exercise' in col_name and 'price' in col_name:
                    headers['exercise_price'] = col_idx
                elif col_name in ['id', 'stakeholder id']:
                    headers['id'] = col_idx
                elif 'name' in col_name or 'optionholder' in col_name:
                    headers['name'] = col_idx
            break
    
    if not header_row_idx or 'options_outstanding' not in headers or 'exercise_price' not in headers:
        return []
    
    # Extract option data
    for row in sheet.iter_rows(min_row=header_row_idx + 1, values_only=True):
        try:
            outstanding = row[headers['options_outstanding']]
            exercise_price = row[headers['exercise_price']]
            
            # Skip rows with no outstanding options
            if not outstanding or not exercise_price:
                continue
            
            # Must be a positive number
            if isinstance(outstanding, (int, float)) and outstanding > 0:
                # Clean exercise price (handle $ and commas)
                price_str = str(exercise_price).replace('$', '').replace(',', '').strip()
                price_float = float(price_str)
                
                option_entry = {
                    'options_outstanding': float(outstanding),
                    'exercise_price': price_float
                }
                
                # Add optional identifying info
                if 'id' in headers and headers['id'] < len(row):
                    option_entry['id'] = row[headers['id']]
                if 'name' in headers and headers['name'] < len(row):
                    option_entry['name'] = row[headers['name']]
                
                options.append(option_entry)
        
        except (ValueError, TypeError, AttributeError, IndexError):
            continue
    
    return options

# System prompt for OpenAI
EXTRACTION_SYSTEM_PROMPT = """You are an expert financial analyst specializing in venture capital cap tables.

You receive STRUCTURED data extracted from Excel cap tables.

INPUT FORMAT:
{
  "cap_table": {
    "totals": {"Common Stock": 8054469, "Series Seed Preferred": 2285713, ...},
    "prices": {"Series Seed Preferred": 0.44, "Series A Preferred Stock": 42.57, ...},
    "options_outstanding": 899337,
    "option_pool_available": 1167233
  },
  "option_ledger": [
    {"options_outstanding": 114286, "exercise_price": 0.81},
    ...
  ]
}

INSTRUCTIONS:

1. **Create security entries:**
   - For each key in "totals", create a security
   - shares_outstanding = value from "totals"
   - original_investment_per_share = value from "prices" (use 0.0 if missing)

2. **Process options:**
   - Group "option_ledger" by "exercise_price"
   - Sum "options_outstanding" for each price
   - Create entries: "Options at $X.XX Exercise Price"
   - Verify total ≈ cap_table["options_outstanding"]

3. **Set total_option_pool_shares:**
   - Use cap_table["option_pool_available"]

4. **Seniority:**
   - All preferred with 1.0x liquidation = seniority 1
   - Common/Options = seniority null

5. **Defaults:**
   - Preferred: liquidation_preference_multiple = 1.0
   - Common/Options: liquidation_preference_multiple = 0.0
   - All: is_participating = false, participation_cap_multiple = 0.0, cumulative_dividend_rate = 0.0, years_since_issuance = 0.0

Return ONLY valid JSON (no markdown):
{
  "securities": [...],
  "total_option_pool_shares": 1167233
}"""

# API endpoints
@app.get("/")
async def root():
    return {"message": "Capital Structure API", "version": "2.0.0"}

@app.get("/health")
async def health():
    return {"status": "healthy"}

@app.post("/api/documents/upload", response_model=DocumentUploadResponse)
async def upload_document(request: FileUploadRequest):
    """Upload a document for processing"""
    try:
        file_bytes = base64.b64decode(request.file_content)
        file_extension = request.file_name.split('.')[-1].lower()
        file_id = f"upload_{uuid.uuid4()}.{file_extension}"
        
        file_storage[file_id] = {
            "content": file_bytes,
            "original_name": request.file_name,
            "size": len(file_bytes),
            "extension": file_extension
        }
        
        return DocumentUploadResponse(
            file_id=file_id,
            file_name=request.file_name,
            message="File uploaded successfully",
            file_size_bytes=len(file_bytes)
        )
        
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Upload failed: {str(e)}")

@app.post("/api/documents/extract", response_model=CapitalStructureInput)
async def extract_data(request: DocumentExtractRequest):
    """Extract capital structure data from uploaded document"""
    try:
        print(f"Extract requested for file_id: {request.file_id}")
        
        if request.file_id not in file_storage:
            raise HTTPException(status_code=404, detail="File not found")
        
        file_data = file_storage[request.file_id]
        file_bytes = file_data["content"]
        file_extension = file_data.get("extension", "txt")
        
        print(f"Processing file: {file_data['original_name']} (type: {file_extension})")
        
        # Process based on file type
        if file_extension in ['xlsx', 'xls']:
            try:
                print("Parsing Excel file...")
                structured_data = parse_excel_cap_table(file_bytes)
                
                # Debug logging
                print("=" * 60)
                print("EXTRACTED CAP TABLE DATA:")
                print("=" * 60)
                cap_table = structured_data.get("cap_table", {})
                print(f"Headers found: {cap_table.get('headers', [])}")
                print(f"Totals: {json.dumps(cap_table.get('totals', {}), indent=2)}")
                print(f"Prices: {json.dumps(cap_table.get('prices', {}), indent=2)}")
                print(f"Options outstanding: {cap_table.get('options_outstanding')}")
                print(f"Option pool available: {cap_table.get('option_pool_available')}")
                print(f"\nOption ledger entries: {len(structured_data.get('option_ledger', []))}")
                if structured_data.get('option_ledger'):
                    # Group by exercise price for summary
                    by_price = defaultdict(list)
                    for opt in structured_data['option_ledger']:
                        by_price[opt['exercise_price']].append(opt['options_outstanding'])
                    for price, outstandings in sorted(by_price.items()):
                        total = sum(outstandings)
                        print(f"  ${price}: {len(outstandings)} grants = {total:,.0f} shares")
                print("=" * 60)
                
                document_text = json.dumps(structured_data, indent=2)
            except Exception as excel_error:
                print(f"ERROR parsing Excel: {str(excel_error)}")
                import traceback
                traceback.print_exc()
                raise HTTPException(status_code=500, detail=f"Excel parsing failed: {str(excel_error)}")
        else:
            print("Processing as text file...")
            document_text = file_bytes.decode("utf-8")
        
        # Call OpenAI
        try:
            print("Calling OpenAI API...")
            response = openai.ChatCompletion.create(
                model="gpt-4o-mini",
                messages=[
                    {"role": "system", "content": EXTRACTION_SYSTEM_PROMPT},
                    {"role": "user", "content": f"Extract capital structure:\n\n{document_text}"}
                ],
                temperature=0.1,
                max_tokens=3000
            )
            
            extracted_json = response.choices[0].message.content.strip()
            print(f"OpenAI response received ({len(extracted_json)} chars)")
            
            # Remove markdown if present
            if extracted_json.startswith("```"):
                extracted_json = extracted_json.split("```")[1]
                if extracted_json.startswith("json"):
                    extracted_json = extracted_json[4:]
                extracted_json = extracted_json.strip()
            
            print("Validating response...")
            result = CapitalStructureInput.model_validate_json(extracted_json)
            print("Validation successful!")
            return result
            
        except openai.error.OpenAIError as openai_error:
            print(f"OpenAI API error: {str(openai_error)}")
            raise HTTPException(status_code=500, detail=f"OpenAI API error: {str(openai_error)}")
        except json.JSONDecodeError as json_error:
            print(f"JSON parsing error: {str(json_error)}")
            print(f"AI Response: {extracted_json[:500]}...")
            raise HTTPException(status_code=500, detail=f"Failed to parse AI response: {str(json_error)}")
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"Unexpected error in extract_data: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Extraction failed: {str(e)}")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=int(os.getenv("PORT", 8000))), '').replace(',', '').strip()
                try:
                    price = float(price_str)
                    if price >= 0:
                        print(f"      {header}: ${price}")
                        data["prices"][header] = price
                except (ValueError, TypeError):
                    pass
        
        # Extract options outstanding
        elif 'option' in row_label and 'outstanding' in row_label and 'issued' in row_label:
            for cell_value in row[1:]:
                if cell_value and isinstance(cell_value, (int, float)) and cell_value > 0:
                    data["options_outstanding"] = float(cell_value)
                    break
        
        # Extract option pool available
        elif 'available for issuance' in row_label or ('shares available' in row_label and 'plan' in row_label):
            for cell_value in row[1:]:
                if cell_value and isinstance(cell_value, (int, float)) and cell_value > 0:
                    data["option_pool_available"] = float(cell_value)
                    break
    
    return data

def extract_option_ledger(sheet) -> List[Dict[str, Any]]:
    """Extract option grants from option ledger sheet"""
    options = []
    
    # Find header row
    header_row_idx = None
    headers = {}
    
    for idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=15, values_only=True)):
        row_text = ' '.join([str(cell or '').lower() for cell in row])
        
        # Look for key indicators of header row
        if 'exercise' in row_text and 'price' in row_text:
            header_row_idx = idx + 1
            
            # Map column names to indices
            for col_idx, cell in enumerate(row):
                if not cell:
                    continue
                
                col_name = str(cell).lower().strip()
                
                # Look for key columns
                if 'outstanding' in col_name and 'option' in col_name:
                    headers['options_outstanding'] = col_idx
                elif 'granted' in col_name and 'option' in col_name:
                    headers['options_granted'] = col_idx
                elif 'exercise' in col_name and 'price' in col_name:
                    headers['exercise_price'] = col_idx
                elif col_name in ['id', 'stakeholder id']:
                    headers['id'] = col_idx
                elif 'name' in col_name or 'optionholder' in col_name:
                    headers['name'] = col_idx
            break
    
    if not header_row_idx or 'options_outstanding' not in headers or 'exercise_price' not in headers:
        return []
    
    # Extract option data
    for row in sheet.iter_rows(min_row=header_row_idx + 1, values_only=True):
        try:
            outstanding = row[headers['options_outstanding']]
            exercise_price = row[headers['exercise_price']]
            
            # Skip rows with no outstanding options
            if not outstanding or not exercise_price:
                continue
            
            # Must be a positive number
            if isinstance(outstanding, (int, float)) and outstanding > 0:
                # Clean exercise price (handle $ and commas)
                price_str = str(exercise_price).replace('$', '').replace(',', '').strip()
                price_float = float(price_str)
                
                option_entry = {
                    'options_outstanding': float(outstanding),
                    'exercise_price': price_float
                }
                
                # Add optional identifying info
                if 'id' in headers and headers['id'] < len(row):
                    option_entry['id'] = row[headers['id']]
                if 'name' in headers and headers['name'] < len(row):
                    option_entry['name'] = row[headers['name']]
                
                options.append(option_entry)
        
        except (ValueError, TypeError, AttributeError, IndexError):
            continue
    
    return options

# System prompt for OpenAI
EXTRACTION_SYSTEM_PROMPT = """You are an expert financial analyst specializing in venture capital cap tables.

You receive STRUCTURED data extracted from Excel cap tables.

INPUT FORMAT:
{
  "cap_table": {
    "totals": {"Common Stock": 8054469, "Series Seed Preferred": 2285713, ...},
    "prices": {"Series Seed Preferred": 0.44, "Series A Preferred Stock": 42.57, ...},
    "options_outstanding": 899337,
    "option_pool_available": 1167233
  },
  "option_ledger": [
    {"options_outstanding": 114286, "exercise_price": 0.81},
    ...
  ]
}

INSTRUCTIONS:

1. **Create security entries:**
   - For each key in "totals", create a security
   - shares_outstanding = value from "totals"
   - original_investment_per_share = value from "prices" (use 0.0 if missing)

2. **Process options:**
   - Group "option_ledger" by "exercise_price"
   - Sum "options_outstanding" for each price
   - Create entries: "Options at $X.XX Exercise Price"
   - Verify total ≈ cap_table["options_outstanding"]

3. **Set total_option_pool_shares:**
   - Use cap_table["option_pool_available"]

4. **Seniority:**
   - All preferred with 1.0x liquidation = seniority 1
   - Common/Options = seniority null

5. **Defaults:**
   - Preferred: liquidation_preference_multiple = 1.0
   - Common/Options: liquidation_preference_multiple = 0.0
   - All: is_participating = false, participation_cap_multiple = 0.0, cumulative_dividend_rate = 0.0, years_since_issuance = 0.0

Return ONLY valid JSON (no markdown):
{
  "securities": [...],
  "total_option_pool_shares": 1167233
}"""

# API endpoints
@app.get("/")
async def root():
    return {"message": "Capital Structure API", "version": "2.0.0"}

@app.get("/health")
async def health():
    return {"status": "healthy"}

@app.post("/api/documents/upload", response_model=DocumentUploadResponse)
async def upload_document(request: FileUploadRequest):
    """Upload a document for processing"""
    try:
        file_bytes = base64.b64decode(request.file_content)
        file_extension = request.file_name.split('.')[-1].lower()
        file_id = f"upload_{uuid.uuid4()}.{file_extension}"
        
        file_storage[file_id] = {
            "content": file_bytes,
            "original_name": request.file_name,
            "size": len(file_bytes),
            "extension": file_extension
        }
        
        return DocumentUploadResponse(
            file_id=file_id,
            file_name=request.file_name,
            message="File uploaded successfully",
            file_size_bytes=len(file_bytes)
        )
        
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Upload failed: {str(e)}")

@app.post("/api/documents/extract", response_model=CapitalStructureInput)
async def extract_data(request: DocumentExtractRequest):
    """Extract capital structure data from uploaded document"""
    try:
        print(f"Extract requested for file_id: {request.file_id}")
        
        if request.file_id not in file_storage:
            raise HTTPException(status_code=404, detail="File not found")
        
        file_data = file_storage[request.file_id]
        file_bytes = file_data["content"]
        file_extension = file_data.get("extension", "txt")
        
        print(f"Processing file: {file_data['original_name']} (type: {file_extension})")
        
        # Process based on file type
        if file_extension in ['xlsx', 'xls']:
            try:
                print("Parsing Excel file...")
                structured_data = parse_excel_cap_table(file_bytes)
                
                # Debug logging
                print("=" * 60)
                print("EXTRACTED CAP TABLE DATA:")
                print("=" * 60)
                cap_table = structured_data.get("cap_table", {})
                print(f"Headers found: {cap_table.get('headers', [])}")
                print(f"Totals: {json.dumps(cap_table.get('totals', {}), indent=2)}")
                print(f"Prices: {json.dumps(cap_table.get('prices', {}), indent=2)}")
                print(f"Options outstanding: {cap_table.get('options_outstanding')}")
                print(f"Option pool available: {cap_table.get('option_pool_available')}")
                print(f"\nOption ledger entries: {len(structured_data.get('option_ledger', []))}")
                if structured_data.get('option_ledger'):
                    # Group by exercise price for summary
                    by_price = defaultdict(list)
                    for opt in structured_data['option_ledger']:
                        by_price[opt['exercise_price']].append(opt['options_outstanding'])
                    for price, outstandings in sorted(by_price.items()):
                        total = sum(outstandings)
                        print(f"  ${price}: {len(outstandings)} grants = {total:,.0f} shares")
                print("=" * 60)
                
                document_text = json.dumps(structured_data, indent=2)
            except Exception as excel_error:
                print(f"ERROR parsing Excel: {str(excel_error)}")
                import traceback
                traceback.print_exc()
                raise HTTPException(status_code=500, detail=f"Excel parsing failed: {str(excel_error)}")
        else:
            print("Processing as text file...")
            document_text = file_bytes.decode("utf-8")
        
        # Call OpenAI
        try:
            print("Calling OpenAI API...")
            response = openai.ChatCompletion.create(
                model="gpt-4o-mini",
                messages=[
                    {"role": "system", "content": EXTRACTION_SYSTEM_PROMPT},
                    {"role": "user", "content": f"Extract capital structure:\n\n{document_text}"}
                ],
                temperature=0.1,
                max_tokens=3000
            )
            
            extracted_json = response.choices[0].message.content.strip()
            print(f"OpenAI response received ({len(extracted_json)} chars)")
            
            # Remove markdown if present
            if extracted_json.startswith("```"):
                extracted_json = extracted_json.split("```")[1]
                if extracted_json.startswith("json"):
                    extracted_json = extracted_json[4:]
                extracted_json = extracted_json.strip()
            
            print("Validating response...")
            result = CapitalStructureInput.model_validate_json(extracted_json)
            print("Validation successful!")
            return result
            
        except openai.error.OpenAIError as openai_error:
            print(f"OpenAI API error: {str(openai_error)}")
            raise HTTPException(status_code=500, detail=f"OpenAI API error: {str(openai_error)}")
        except json.JSONDecodeError as json_error:
            print(f"JSON parsing error: {str(json_error)}")
            print(f"AI Response: {extracted_json[:500]}...")
            raise HTTPException(status_code=500, detail=f"Failed to parse AI response: {str(json_error)}")
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"Unexpected error in extract_data: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Extraction failed: {str(e)}")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=int(os.getenv("PORT", 8000)))
